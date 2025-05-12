# this script takes the compiled excel file (with new tabs for each sample) from the writeexcel.py script and creates a new spreadsheet which has columns Sheet_Name (sampleID), Gene_Name (the name of the antibiotic resistance gene), Gene_Function (the function of the gene), and Occurances (how many times that gene was found in that sample). This consolidated sheet was used for analysis in R. 

import xlsxwriter
import openpyxl
from collections import defaultdict
from openpyxl.utils import get_column_letter
import os

# Set the working directory (where compiled data sheet from writeexcel.py script is
os.chdir('/path/to/directory/with/compiled/data')
# Load the source Excel file
src_wb = openpyxl.load_workbook('compiled_data.xlsx')

# Create a new target Excel file 
tgt_wb = xlsxwriter.Workbook('consolidated_data.xlsx')
tgt_ws = tgt_wb.add_worksheet()

# Set the header row
tgt_ws.write('A1', 'Sheet_Name')
tgt_ws.write('B1', 'Gene_Name')
tgt_ws.write('C1', 'Gene_Function')
tgt_ws.write('D1', 'Occurrences')

# Iterate over each sheet in the source 
row_num = 2
for sheet_name in src_wb.sheetnames:
    src_ws = src_wb[sheet_name]
    
    # Create a dictionary to store the occurrences of each gene name and function
    occurrences = defaultdict(int)
    
    # Iterate over each row in the source sheet
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        gene_name = row[5]  # Column 6 - MAY NEED TO ALTER THIS TO FIT YOUR DATA - change 5 to the column number - 1 with the gene names in it
        gene_function = row[14]  # Column 15 - MAY NEED TO ALTER THIS TO FIT YOUR DATA - change 15 to the column number - 1 with the gene functions in it
        
        occurrences[(gene_name, gene_function)] += 1

    # Write the data to the target sheet
    for (gene_name, gene_function), count in occurrences.items():
        if gene_name is not None and gene_name != 'N/A':
            tgt_ws.write(row_num, 0, sheet_name)  
            tgt_ws.write(row_num, 1, gene_name)  
            tgt_ws.write(row_num, 2, gene_function)  
            tgt_ws.write(row_num, 3, count)  
            row_num += 1
                            
# Close the target Excel file
tgt_wb.close()
